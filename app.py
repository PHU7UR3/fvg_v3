import os
import re
import time
import json
import logging
import threading
import requests as req
from datetime import datetime, timedelta, timezone
from flask import Flask, jsonify, render_template_string, request, send_file

import alpaca_trade_api as tradeapi
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
API_KEY       = os.environ.get("API_KEY", "")
SECRET_KEY    = os.environ.get("SECRET_KEY", "")
BASE_URL      = re.sub(r'/v2/?$', '', os.environ.get("BASE_URL", "https://paper-api.alpaca.markets")).rstrip("/")
WATCHLIST_ENV = os.environ.get("WATCHLIST", "NVDA,AMD,ASML,AMAT,LRCX,TSM,XOM,RTX,UNH,JPM")
TIMEFRAME     = os.environ.get("TIMEFRAME", "5Min")
STATE_FILE    = "/tmp/fvg_state.json"
EXCEL_FILE    = "/tmp/fvg_trades.xlsx"

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(message)s")
log = logging.getLogger(__name__)
lock = threading.Lock()

# ─────────────────────────────────────────────
# STATE
# ─────────────────────────────────────────────
def default_state():
    return {
        "watchlist": [w.strip() for w in WATCHLIST_ENV.split(",") if w.strip()],
        "logs": [], "trades": [],
        "fvg_count": 0, "wins": 0, "losses": 0, "total_pnl": 0.0,
        "settings": {
            "timeframe": TIMEFRAME,
            "fvg_min_size": 0.001,
            "risk_per_trade": 0.02,
            "reward_ratio": 2.0,
            "max_positions": 3,
            "check_interval": 60,
            "use_rsi": True,
            "use_volume": False,
            "use_ema_trend": True,
            "rsi_period": 14,
            "rsi_oversold": 20,
            "rsi_overbought": 80,
        }
    }

def load_state():
    try:
        if os.path.exists(STATE_FILE):
            with open(STATE_FILE) as f:
                saved = json.load(f)
                base = default_state()
                base.update(saved)
                for k, v in default_state()["settings"].items():
                    if k not in base["settings"]:
                        base["settings"][k] = v
                return base
    except:
        pass
    return default_state()

def save_state():
    try:
        with open(STATE_FILE, "w") as f:
            json.dump({k: state[k] for k in
                ["watchlist","logs","trades","fvg_count","wins","losses","total_pnl","settings"]
            }, f)
    except Exception as e:
        log.error(f"Save: {e}")

runtime = {
    "running": False, "market_open": False,
    "equity": 0.0, "cash": 0.0,
    "positions": [], "last_scan": "Never", "next_open": "",
}
state = load_state()

def add_log(msg, level="info"):
    with lock:
        state["logs"].insert(0, {"time": datetime.now().strftime("%H:%M:%S"), "msg": msg, "level": level})
        if len(state["logs"]) > 100: state["logs"] = state["logs"][:100]
    save_state()
    log.info(msg)

# ─────────────────────────────────────────────
# EXCEL
# ─────────────────────────────────────────────
def init_excel():
    if os.path.exists(EXCEL_FILE): return
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Trade Log"
    headers = ["Date","Time","Symbol","Side","Qty","Entry","SL","TP","Exit","P&L ($)","P&L (%)","Status","FVG","Trend","RSI","Notes"]
    hf = PatternFill("solid", fgColor="0D0D14")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="00FF88", name="Arial", size=10)
        c.fill = hf; c.alignment = Alignment(horizontal="center")
    for i,w in enumerate([12,10,8,6,5,10,10,10,10,10,8,8,8,8,6,30],1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "FVG Bot Trade Summary"
    ws2["A1"].font = Font(bold=True, size=14, color="00FF88")
    for i,l in enumerate(["Total","Wins","Losses","Win%","Total P&L","Avg P&L","Best","Worst"],3):
        ws2.cell(row=i,column=1,value=l).font = Font(bold=True)
    ws2["B3"]="=COUNTA('Trade Log'!A2:A9999)-1"
    ws2["B4"]="=COUNTIF('Trade Log'!L2:L9999,\"WIN\")"
    ws2["B5"]="=COUNTIF('Trade Log'!L2:L9999,\"LOSS\")"
    ws2["B6"]="=IFERROR(B4/B3*100,0)"
    ws2["B7"]="=IFERROR(SUM('Trade Log'!J2:J9999),0)"
    ws2["B8"]="=IFERROR(AVERAGE('Trade Log'!J2:J9999),0)"
    ws2["B9"]="=IFERROR(MAX('Trade Log'!J2:J9999),0)"
    ws2["B10"]="=IFERROR(MIN('Trade Log'!J2:J9999),0)"
    wb.save(EXCEL_FILE)

def _write_row(ws, nr, t):
    pnl = t.get("pnl",0.0); ent = t.get("entry",1); qty = t.get("qty",1)
    pct = round((pnl/(ent*qty))*100,2) if ent and qty else 0
    st  = t.get("status","OPEN")
    row = [t.get("date",""),t.get("time",""),t.get("symbol",""),t.get("side","").upper(),
           qty,ent,t.get("sl",0),t.get("tp",0),t.get("exit_price","") or "",
           round(pnl,2),pct,st,t.get("fvg_type",""),t.get("trend",""),t.get("rsi",""),t.get("notes","")]
    for col,val in enumerate(row,1):
        cell = ws.cell(row=nr,column=col,value=val)
        cell.font = Font(name="Arial",size=9)
        cell.alignment = Alignment(horizontal="center")
        if col==10 and isinstance(val,(int,float)):
            cell.font = Font(name="Arial",size=9,color="00AA44" if val>0 else "CC0000" if val<0 else "888888")
        if col==12:
            cl = {"WIN":("002200","00FF88"),"LOSS":("220000","FF3D6E")}.get(val,("1a1a2e","FFD166"))
            cell.fill = PatternFill("solid",fgColor=cl[0])
            cell.font = Font(name="Arial",size=9,color=cl[1],bold=True)
        if col==4:
            cell.font = Font(name="Arial",size=9,color="00FF88" if val=="BUY" else "FF3D6E",bold=True)

def log_excel(trade):
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE); ws = wb["Trade Log"]
        nr = trade.get("excel_row")
        if not nr:
            for r in range(2, ws.max_row+1):
                if ws.cell(r,3).value==trade.get("symbol") and ws.cell(r,2).value==trade.get("time"):
                    nr = r; break
        if not nr: nr = ws.max_row+1
        trade["excel_row"] = nr
        _write_row(ws, nr, trade)
        wb.save(EXCEL_FILE)
    except Exception as e:
        log.error(f"Excel: {e}")

# ─────────────────────────────────────────────
# ALPACA HELPERS
# ─────────────────────────────────────────────
def get_api():
    return tradeapi.REST(API_KEY, SECRET_KEY, BASE_URL, api_version="v2")

def alpaca_headers():
    return {"APCA-API-KEY-ID": API_KEY, "APCA-API-SECRET-KEY": SECRET_KEY}

def setup_account():
    """Disable PDT, cancel all orders, close all positions on startup."""
    try:
        r = req.patch(f"{BASE_URL}/v2/account/configurations",
                      headers=alpaca_headers(), json={"pdt_check": "entry"})
        add_log(f"✅ PDT disabled ({r.status_code})")
    except Exception as e:
        add_log(f"⚠️ PDT: {e}", "error")

    try:
        api = get_api()
        orders = api.list_orders(status="open")
        if orders:
            api.cancel_all_orders()
            add_log(f"🗑️ Cancelled {len(orders)} open orders")
    except Exception as e:
        add_log(f"⚠️ Cancel orders: {e}", "error")

    try:
        api = get_api()
        positions = api.list_positions()
        if positions:
            for p in positions:
                try: api.close_position(p.symbol)
                except: pass
            add_log(f"🔴 Closed {len(positions)} existing positions — fresh start")
            time.sleep(2)
    except Exception as e:
        add_log(f"⚠️ Close positions: {e}", "error")

def is_market_open():
    utc = datetime.now(timezone.utc)
    mins = utc.hour*60 + utc.minute
    weekday = utc.weekday() < 5
    manual = weekday and 810 <= mins <= 1200
    try:
        api = get_api()
        clock = api.get_clock()
        with lock: runtime["next_open"] = str(clock.next_open)[:16]
        return clock.is_open or manual
    except:
        return manual

# ─────────────────────────────────────────────
# INDICATORS
# ─────────────────────────────────────────────
def get_candles(symbol, tf=None, hours=10):
    try:
        api = get_api()
        tf  = tf or state["settings"]["timeframe"]
        end = datetime.utcnow()
        bars = api.get_bars(symbol, tf,
            start=(end-timedelta(hours=hours)).isoformat()+"Z",
            end=end.isoformat()+"Z", limit=100, feed="iex").df
        if hasattr(bars.index,"tz") and bars.index.tz:
            bars.index = bars.index.tz_localize(None)
        return bars if len(bars)>=5 else None
    except Exception as e:
        err = str(e).lower()
        if not any(x in err for x in ["subscription","iex","forbidden","market","closed"]):
            add_log(f"⚠️ Candle {symbol}: {e}", "error")
        return None

def calc_rsi(closes, period=14):
    try:
        if len(closes)<period+1: return 50.0
        d=closes.diff(); g=d.where(d>0,0.0).rolling(period).mean()
        l=(-d.where(d<0,0.0)).rolling(period).mean()
        v=float((100-(100/(1+(g/l)))).iloc[-1])
        return round(v,1) if not pd.isna(v) else 50.0
    except: return 50.0

def get_trend(symbol):
    try:
        bars = get_candles(symbol, tf="1Hour", hours=72)
        if bars is None or len(bars)<21: return "neutral"
        c=bars["close"]; p=c.iloc[-1]
        e20=c.ewm(span=20,adjust=False).mean().iloc[-1]
        e50=c.ewm(span=min(50,len(c)),adjust=False).mean().iloc[-1]
        if p>e20 and e20>e50*0.999: return "bullish"
        if p<e20 and e20<e50*1.001: return "bearish"
        return "neutral"
    except: return "neutral"

def detect_fvg(bars):
    fvgs=[]; ms=state["settings"]["fvg_min_size"]
    for i in range(2,len(bars)):
        c1,c2,c3=bars.iloc[i-2],bars.iloc[i-1],bars.iloc[i]
        body=abs(float(c2["close"])-float(c2["open"]))
        rng=float(c2["high"])-float(c2["low"])
        if rng==0 or body/rng<0.3: continue
        if float(c3["low"])>float(c1["high"]):
            gap=(float(c3["low"])-float(c1["high"]))/float(c1["high"])
            if gap>=ms and float(c2["close"])>float(c2["open"]):
                fvgs.append({"type":"bullish","top":float(c3["low"]),"bottom":float(c1["high"]),
                             "gap_size":round(gap*100,3),"score":gap*100*(i/len(bars)),
                             "c2_vol":float(c2.get("volume",0))})
        elif float(c3["high"])<float(c1["low"]):
            gap=(float(c1["low"])-float(c3["high"]))/float(c1["low"])
            if gap>=ms and float(c2["close"])<float(c2["open"]):
                fvgs.append({"type":"bearish","top":float(c1["low"]),"bottom":float(c3["high"]),
                             "gap_size":round(gap*100,3),"score":gap*100*(i/len(bars)),
                             "c2_vol":float(c2.get("volume",0))})
    return sorted(fvgs,key=lambda x:x["score"],reverse=True)

def price_in_fvg(price, fvg):
    tol = fvg["bottom"]*0.0005
    return (fvg["bottom"]-tol) <= price <= fvg["top"]

def calc_qty(equity, cash, entry, sl):
    risk=state["settings"]["risk_per_trade"]
    per=abs(entry-sl)
    if per==0: return 0
    by_risk=(equity*risk)/per
    by_cash=(cash*0.9)/entry
    by_eq=(equity*0.25)/entry
    qty=int(min(by_risk,by_cash,by_eq))
    return max(1,qty) if cash>=entry else 0

# ─────────────────────────────────────────────
# TRADE EXECUTION
# ─────────────────────────────────────────────
def place_order(symbol, side, qty, entry, sl, tp, fvg, trend, rsi):
    try:
        api = get_api()
        entry=round(float(entry),2); sl=round(float(sl),2); tp=round(float(tp),2)
        # Fix invalid SL/TP
        if side=="buy":
            if sl>=entry: sl=round(entry*0.985,2)
            if tp<=entry: tp=round(entry*1.03,2)
        else:
            if sl<=entry: sl=round(entry*1.015,2)
            if tp>=entry: tp=round(entry*0.97,2)

        otype = "BRACKET"
        try:
            api.submit_order(symbol=symbol,qty=qty,side=side,
                type="limit",time_in_force="gtc",limit_price=entry,
                order_class="bracket",
                stop_loss={"stop_price":sl},
                take_profit={"limit_price":tp})
        except Exception as be:
            if "pattern day" in str(be).lower() or "pdt" in str(be).lower():
                api.submit_order(symbol=symbol,qty=qty,side=side,
                    type="limit",time_in_force="gtc",limit_price=entry)
                otype="LIMIT"
            else:
                raise be

        add_log(f"✅ {side.upper()} {qty} {symbol} @${entry} SL:${sl} TP:${tp} [{otype}]","trade")
        trade = {
            "time":datetime.now().strftime("%H:%M:%S"),
            "date":datetime.now().strftime("%d %b %Y"),
            "symbol":symbol,"side":side,"qty":qty,
            "entry":entry,"sl":sl,"tp":tp,
            "exit_price":None,"pnl":0.0,"status":"OPEN",
            "fvg_type":fvg["type"],"gap_size":fvg["gap_size"],
            "trend":trend,"rsi":rsi,"excel_row":None,
            "notes":f"FVG {fvg['gap_size']}% {trend} RSI={rsi} [{otype}]"
        }
        with lock:
            state["trades"].insert(0,trade)
            if len(state["trades"])>100: state["trades"]=state["trades"][:100]
        log_excel(trade)
        save_state()
    except Exception as e:
        add_log(f"❌ Order {symbol}: {e}","error")

def monitor_and_close():
    """Close positions that hit TP or SL."""
    try:
        api = get_api()
        positions = api.list_positions()
        if not positions: return
        for pos in positions:
            sym     = pos.symbol
            current = float(pos.current_price)
            qty     = abs(int(float(pos.qty)))
            is_long = float(pos.qty) > 0
            with lock:
                trade = next((t for t in state["trades"]
                              if t["symbol"]==sym and t["status"]=="OPEN"), None)
            if not trade: continue
            tp=trade.get("tp",0); sl=trade.get("sl",0)
            hit = None
            if is_long:
                if tp and current>=tp: hit="TP"
                elif sl and current<=sl: hit="SL"
            else:
                if tp and current<=tp: hit="TP"
                elif sl and current>=sl: hit="SL"
            if hit:
                try:
                    close_side = "sell" if is_long else "buy"
                    api.submit_order(symbol=sym,qty=qty,side=close_side,
                                     type="market",time_in_force="gtc")
                    pnl = float(pos.unrealized_pl)
                    icon = "✅" if pnl>0 else "❌"
                    add_log(f"{icon} CLOSED {sym} [{hit}] @${current:.2f} PnL:${pnl:.2f}","trade")
                    with lock:
                        for t in state["trades"]:
                            if t["symbol"]==sym and t["status"]=="OPEN":
                                t.update({"pnl":round(pnl,2),"exit_price":current,
                                          "status":"WIN" if pnl>0 else "LOSS"})
                                log_excel(t); break
                        if pnl>0: state["wins"]+=1
                        else: state["losses"]+=1
                        state["total_pnl"]=round(state.get("total_pnl",0)+pnl,2)
                    save_state()
                except Exception as e:
                    add_log(f"⚠️ Close {sym}: {e}","error")
    except Exception as e:
        log.error(f"Monitor: {e}")

def update_pnl():
    """Update PnL from closed Alpaca orders."""
    try:
        api = get_api()
        orders = api.list_orders(status="closed",limit=50)
        total=wins=losses=0
        with lock:
            for t in state["trades"]:
                if t["status"]!="OPEN":
                    total+=t["pnl"]
                    if t["pnl"]>0: wins+=1
                    elif t["pnl"]<0: losses+=1
                    continue
                for o in orders:
                    if o.symbol!=t["symbol"] or not o.filled_avg_price: continue
                    fp=float(o.filled_avg_price); fq=float(o.filled_qty or 0)
                    os_=str(o.side).lower()
                    if (t["side"]=="buy" and os_=="sell") or (t["side"]=="sell" and os_=="buy"):
                        pnl=(fp-t["entry"])*fq if t["side"]=="buy" else (t["entry"]-fp)*fq
                        t.update({"pnl":round(pnl,2),"exit_price":fp,
                                  "status":"WIN" if pnl>0 else "LOSS"})
                        total+=pnl
                        if pnl>0: wins+=1
                        else: losses+=1
                        log_excel(t); break
            state.update({"total_pnl":round(total,2),"wins":wins,"losses":losses})
        save_state()
    except Exception as e:
        log.error(f"PnL update: {e}")

# ─────────────────────────────────────────────
# MAIN BOT LOOP
# ─────────────────────────────────────────────
def bot_loop():
    add_log("🤖 FVG Bot started")
    try:
        api = get_api()
        acc = api.get_account()
        with lock:
            runtime["equity"]=float(acc.equity)
            runtime["cash"]=float(acc.cash)
        add_log(f"💰 Connected: ${float(acc.equity):,.2f}")
    except Exception as e:
        add_log(f"❌ Connection failed: {e}","error")
        with lock: runtime["running"]=False
        return

    setup_account()

    while True:
        with lock:
            if not runtime["running"]: break
        try:
            # Refresh account
            try:
                api=get_api(); acc=api.get_account()
                with lock:
                    runtime["equity"]=float(acc.equity)
                    runtime["cash"]=float(acc.cash)
            except: pass

            # Market check
            open_ = is_market_open()
            with lock: runtime["market_open"]=open_

            if not open_:
                add_log(f"💤 Market closed — next: {runtime['next_open']} UTC")
                time.sleep(60); continue

            # Get real positions from Alpaca
            api = get_api()
            positions = api.list_positions()
            pos_dict = {p.symbol:p for p in positions}
            with lock:
                runtime["last_scan"]=datetime.now().strftime("%H:%M:%S")
                runtime["positions"]=[{
                    "symbol":p.symbol,"qty":p.qty,
                    "entry":float(p.avg_entry_price),"current":float(p.current_price),
                    "pnl":float(p.unrealized_pl),
                    "pnl_pct":round(float(p.unrealized_plpc)*100,2),
                    "side":"long" if float(p.qty)>0 else "short"
                } for p in positions]

            # Monitor and close positions at TP/SL
            monitor_and_close()
            update_pnl()

            s=state["settings"]
            equity=runtime["equity"]; cash=runtime["cash"]
            max_pos=s["max_positions"]

            add_log(f"📡 {len(state['watchlist'])} stocks | ${equity:,.0f} | {len(pos_dict)}/{max_pos} pos")

            if len(pos_dict)>=max_pos:
                add_log(f"⚠️ {max_pos} positions full — monitoring for closes")
                time.sleep(s["check_interval"]); continue

            with lock: watchlist=state["watchlist"][:]

            for symbol in watchlist:
                with lock:
                    if not runtime["running"]: break
                if symbol in pos_dict: continue

                bars=get_candles(symbol)
                if bars is None or len(bars)<10: continue

                price=float(bars["close"].iloc[-1])
                rsi=calc_rsi(bars["close"],int(s["rsi_period"]))
                trend=get_trend(symbol) if s["use_ema_trend"] else "neutral"
                fvgs=detect_fvg(bars)
                if not fvgs: continue

                with lock: state["fvg_count"]+=len(fvgs)

                traded=False
                for fvg in fvgs[:5]:
                    if traded: break
                    if not price_in_fvg(price,fvg): continue
                    reward=s["reward_ratio"]
                    add_log(f"🎯 {symbol} {fvg['type']} gap={fvg['gap_size']}% RSI={rsi} {trend}")

                    if fvg["type"]=="bullish" and trend in ["bullish","neutral"]:
                        if not s["use_rsi"] or rsi<s["rsi_overbought"]:
                            sl=round(fvg["bottom"]*0.997,2)
                            tp=round(price+(price-sl)*reward,2)
                            qty=calc_qty(equity,cash,price,sl)
                            if qty>0:
                                place_order(symbol,"buy",qty,price,sl,tp,fvg,trend,rsi)
                                traded=True

                    elif fvg["type"]=="bearish" and trend in ["bearish","neutral"]:
                        if not s["use_rsi"] or rsi>s["rsi_oversold"]:
                            sl=round(fvg["top"]*1.003,2)
                            tp=round(price-(sl-price)*reward,2)
                            qty=calc_qty(equity,cash,price,sl)
                            if qty>0:
                                place_order(symbol,"sell",qty,price,sl,tp,fvg,trend,rsi)
                                traded=True

            save_state()
            time.sleep(s["check_interval"])

        except Exception as e:
            add_log(f"❌ Error: {e}","error")
            time.sleep(30)

    add_log("⏹ Stopped")

def watchdog():
    time.sleep(20)
    while True:
        time.sleep(30)
        with lock: running=runtime["running"]
        if running:
            alive=any(t.name=="bot" and t.is_alive() for t in threading.enumerate())
            if not alive:
                add_log("🔄 Watchdog restarting bot","error")
                threading.Thread(target=bot_loop,daemon=True,name="bot").start()

def start_bot():
    with lock:
        if runtime["running"]: return
        runtime["running"]=True
    threading.Thread(target=bot_loop,daemon=True,name="bot").start()

# ─────────────────────────────────────────────
# FLASK
# ─────────────────────────────────────────────
app = Flask(__name__)

@app.route("/")
def index():
    with open(os.path.join(os.path.dirname(__file__),"dashboard.html")) as f:
        return render_template_string(f.read())

@app.route("/api/status")
def api_status():
    try:
        acc=get_api().get_account()
        with lock:
            runtime["equity"]=float(acc.equity)
            runtime["cash"]=float(acc.cash)
    except: pass
    with lock:
        return jsonify({
            "running":runtime["running"],"market_open":runtime["market_open"],
            "next_open":runtime["next_open"],"equity":runtime["equity"],
            "cash":runtime["cash"],"positions":runtime["positions"],
            "last_scan":runtime["last_scan"],"fvg_count":state["fvg_count"],
            "total_pnl":state["total_pnl"],"wins":state["wins"],"losses":state["losses"],
            "watchlist":state["watchlist"],"logs":state["logs"][:30],
            "trades":state["trades"][:20],"settings":state["settings"],
        })

@app.route("/api/start",methods=["POST"])
def api_start():
    start_bot(); return jsonify({"ok":True})

@app.route("/api/stop",methods=["POST"])
def api_stop():
    with lock: runtime["running"]=False
    return jsonify({"ok":True})

@app.route("/api/close_all",methods=["POST"])
def api_close_all():
    """Close all positions and cancel all orders."""
    results=[]
    try:
        api=get_api()
        try:
            orders=api.list_orders(status="open")
            if orders: api.cancel_all_orders()
            results.append(f"Cancelled {len(orders)} orders")
        except Exception as e: results.append(f"Orders error: {e}")
        try:
            positions=api.list_positions()
            closed=0
            for p in positions:
                try: api.close_position(p.symbol); closed+=1
                except Exception as e: results.append(f"Close {p.symbol}: {e}")
            results.append(f"Closed {closed} positions")
            with lock: runtime["positions"]=[]
        except Exception as e: results.append(f"Positions error: {e}")
        add_log(f"🔴 Close all: {'; '.join(results)}")
        return jsonify({"ok":True,"msg":"; ".join(results)})
    except Exception as e:
        return jsonify({"ok":False,"error":str(e)})

@app.route("/api/settings",methods=["POST"])
def api_settings():
    data=request.get_json()
    with lock:
        s=state["settings"]
        for key,cast in [("timeframe",str),("fvg_min_size",float),("risk_per_trade",float),
                         ("reward_ratio",float),("max_positions",int),("check_interval",int),
                         ("use_rsi",bool),("use_volume",bool),("use_ema_trend",bool),
                         ("rsi_oversold",float),("rsi_overbought",float)]:
            if key in data: s[key]=cast(data[key])
    save_state()
    add_log(f"⚙️ Settings saved")
    return jsonify({"ok":True,"settings":state["settings"]})

@app.route("/api/watchlist/add/<symbol>",methods=["POST"])
def api_add(symbol):
    symbol=symbol.upper().strip()
    with lock:
        if symbol in state["watchlist"]:
            return jsonify({"ok":False,"msg":f"{symbol} already in list"})
        state["watchlist"].append(symbol)
    save_state(); add_log(f"➕ {symbol}")
    return jsonify({"ok":True,"watchlist":state["watchlist"]})

@app.route("/api/watchlist/remove/<symbol>",methods=["POST"])
def api_remove(symbol):
    symbol=symbol.upper().strip()
    with lock:
        if symbol not in state["watchlist"]:
            return jsonify({"ok":False,"msg":"Not found"})
        state["watchlist"].remove(symbol)
    save_state(); add_log(f"➖ {symbol}")
    return jsonify({"ok":True,"watchlist":state["watchlist"]})

@app.route("/api/clear_logs",methods=["POST"])
def api_clear_logs():
    with lock: state["logs"]=[]
    save_state(); return jsonify({"ok":True})

@app.route("/api/download_excel")
def download_excel():
    if not os.path.exists(EXCEL_FILE): init_excel()
    return send_file(EXCEL_FILE,as_attachment=True,download_name="fvg_trades.xlsx")

@app.route("/ping")
def ping(): return "pong",200

@app.route("/api/test")
def api_test():
    try:
        acc=get_api().get_account()
        return jsonify({"ok":True,"equity":float(acc.equity),"base_url":BASE_URL})
    except Exception as e:
        return jsonify({"ok":False,"error":str(e),"base_url":BASE_URL})

# ─────────────────────────────────────────────
# STARTUP
# ─────────────────────────────────────────────
init_excel()
if API_KEY and SECRET_KEY:
    start_bot()
    threading.Thread(target=watchdog,daemon=True,name="watchdog").start()
else:
    log.warning("⚠️ No API keys")

if __name__=="__main__":
    app.run(host="0.0.0.0",port=int(os.environ.get("PORT",5000)))
